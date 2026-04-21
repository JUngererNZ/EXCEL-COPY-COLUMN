import openpyxl
import re
from openpyxl.utils import get_column_letter
from copy import copy

# Configuration
# Replace with your actual file path
FILENAME = "BARTRAC - KAMOA TRACKING AS OF 31-12-2025.xlsx"
OUTPUT_FILENAME = FILENAME.replace(".xlsx", "_updated.xlsx")
SHEET_NAME = "Sheet1"  # Ensure this matches the sheet name in your Excel file

def process_excel(file_path):
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(file_path)
        
        if SHEET_NAME not in wb.sheetnames:
            print(f"Error: Sheet '{SHEET_NAME}' not found in {file_path}")
            return
            
        ws = wb[SHEET_NAME]

        # 1. Locate the target column
        # Regex pattern: "COMMENTS" followed by space and DD-MM-YYYY format
        pattern = re.compile(r"COMMENTS \d{2}-\d{2}-\d{4}")
        target_col_idx = -1

        # Scan the first row to find the LAST matching header
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value and isinstance(cell_value, str) and pattern.search(cell_value):
                target_col_idx = col
        
        if target_col_idx == -1:
            print("No column matching the pattern 'COMMENTS DD-MM-YYYY' was found.")
            return

        print(f"Target column found at: {get_column_letter(target_col_idx)} (Index {target_col_idx})")

        # 2. Insert a new blank column to the left
        # This shifts the original target column to the right (target_col_idx + 1)
        ws.insert_cols(target_col_idx)
        
        # New column index is target_col_idx, original data is now at target_col_idx + 1
        new_col_idx = target_col_idx
        old_col_idx = target_col_idx + 1
        
        # Copy column width from original to new
        old_col_letter = get_column_letter(old_col_idx)
        new_col_letter = get_column_letter(new_col_idx)
        ws.column_dimensions[new_col_letter].width = ws.column_dimensions[old_col_letter].width

        # 3. Copy cell values and formatting
        # Iterating through all rows in the column
        for row in range(1, ws.max_row + 1):
            src_cell = ws.cell(row=row, column=old_col_idx)
            dest_cell = ws.cell(row=row, column=new_col_idx)
            
            # Copy value
            dest_cell.value = src_cell.value
            
            # Copy style/formatting if the source cell has one
            if src_cell.has_style:
                dest_cell.font = copy(src_cell.font)
                dest_cell.border = copy(src_cell.border)
                dest_cell.fill = copy(src_cell.fill)
                dest_cell.number_format = copy(src_cell.number_format)
                dest_cell.protection = copy(src_cell.protection)
                dest_cell.alignment = copy(src_cell.alignment)

        # 4. Hide the original column
        ws.column_dimensions[old_col_letter].hidden = True
        
        # Save the modified workbook
        wb.save(OUTPUT_FILENAME)
        print(f"Success! Updated file saved as: {OUTPUT_FILENAME}")

    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    process_excel(FILENAME)