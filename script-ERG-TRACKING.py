
# =========================
# Configuration
# input_file = Path(r"C:\Users\Jason\Projects\EXCEL-COPY-COLUMN\BARTRAC - ERG TRACKING 31-03-2026.xlsx")
# sheet_name = "CURRENT SHIPMENTS "
# header_row = 1

import re
from copy import copy
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# =========================
# Configuration
# =========================
input_file = Path(r"C:\Users\Jason\Projects\EXCEL-COPY-COLUMN\BARTRAC - ERG TRACKING 31-03-2026.xlsx")
sheet_name = "CURRENT SHIPMENTS "
header_row = 1

# Header format: COMMENTS DD-MM-YYYY
header_pattern = re.compile(r"^COMMENTS \d{2}-\d{2}-\d{4}$")


# =========================
# Validate workbook path
# =========================
if not input_file.exists():
    raise FileNotFoundError(f"Workbook not found: {input_file}")


# =========================
# Load workbook and worksheet
# =========================
wb = load_workbook(input_file)
if sheet_name not in wb.sheetnames:
    raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")

ws = wb[sheet_name]


# =========================
# Find the LAST matching COMMENTS column in row 6
# =========================
target_col = None

for col_idx in range(1, ws.max_column + 1):
    header_value = ws.cell(row=header_row, column=col_idx).value
    if isinstance(header_value, str) and header_pattern.match(header_value.strip()):
        target_col = col_idx

if target_col is None:
    raise ValueError(
        f"No header matching pattern 'COMMENTS DD-MM-YYYY' found in row {header_row}."
    )


# =========================
# Cache the original column data and formatting
# =========================
column_data = []

for row_idx in range(1, ws.max_row + 1):
    cell = ws.cell(row=row_idx, column=target_col)
    column_data.append({
        "row": row_idx,
        "value": cell.value,
        "font": copy(cell.font),
        "fill": copy(cell.fill),
        "border": copy(cell.border),
        "alignment": copy(cell.alignment),
        "number_format": cell.number_format,
        "protection": copy(cell.protection),
        "comment": copy(cell.comment) if cell.comment else None,
        "hyperlink": copy(cell.hyperlink) if cell.hyperlink else None,
        "has_style": cell.has_style,
    })

# Capture original width
target_letter = get_column_letter(target_col)
original_width = ws.column_dimensions[target_letter].width


# =========================
# Record merged ranges touching the target column
# so we can duplicate them for the new right-hand column
# =========================
merged_ranges_to_duplicate = []

for merged_range in list(ws.merged_cells.ranges):
    if merged_range.min_col <= target_col <= merged_range.max_col:
        width = merged_range.max_col - merged_range.min_col
        offset_inside_merge = target_col - merged_range.min_col

        # Original target stays where it is.
        # New duplicate will be inserted one column to the right.
        new_min_col = target_col + 1 - offset_inside_merge
        new_max_col = new_min_col + width

        merged_ranges_to_duplicate.append((
            merged_range.min_row,
            new_min_col,
            merged_range.max_row,
            new_max_col
        ))


# =========================
# Insert a blank column to the RIGHT of the target column
# =========================
ws.insert_cols(target_col + 1, 1)

# Original target remains at target_col (left side)
# New duplicate goes into target_col + 1 (right side)
new_col = target_col + 1
original_col = target_col


# =========================
# Paste cached data into the new right-hand column
# =========================
for item in column_data:
    row_idx = item["row"]
    new_cell = ws.cell(row=row_idx, column=new_col)

    new_cell.value = item["value"]

    if item["has_style"]:
        new_cell.font = copy(item["font"])
        new_cell.fill = copy(item["fill"])
        new_cell.border = copy(item["border"])
        new_cell.alignment = copy(item["alignment"])
        new_cell.number_format = item["number_format"]
        new_cell.protection = copy(item["protection"])

    if item["comment"] is not None:
        new_cell.comment = item["comment"]

    if item["hyperlink"] is not None:
        new_cell._hyperlink = item["hyperlink"]


# =========================
# Copy column width to the new duplicate column
# =========================
new_letter = get_column_letter(new_col)
ws.column_dimensions[new_letter].width = original_width


# =========================
# Recreate merged cells for the duplicated right-hand column
# Skip conflicts gracefully
# =========================
for min_row, min_col, max_row, max_col in merged_ranges_to_duplicate:
    try:
        ws.merge_cells(
            start_row=min_row,
            start_column=min_col,
            end_row=max_row,
            end_column=max_col
        )
    except ValueError:
        pass


# =========================
# Hide the ORIGINAL column on the LEFT
# =========================
original_letter = get_column_letter(original_col)
ws.column_dimensions[original_letter].hidden = True


# =========================
# Save changes back into the SAME file
# This will overwrite the original workbook
# =========================
wb.save(input_file)

print(f"Done. File updated in place: {input_file}")